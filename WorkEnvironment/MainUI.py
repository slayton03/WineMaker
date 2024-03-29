import tkinter
import tkinter.messagebox
import customtkinter
import os.path
import math
import pandas as pd
# import self as self
from openpyxl import Workbook, load_workbook
from datetime import date

# from pytube import *


if os.path.exists("WineMakerData.xlsx"):
    book = load_workbook("WineMakerData.xlsx")
    current = book.active
    sch = book['Schedule']
    todo = book["Todo"]
    lib = book["Lib"]

else:
    # Create workbook
    book = Workbook()
    current = book.active
    # Create section for wine data
    current.title = "WineData"
    current["A1"] = "Index"
    current["B1"] = "Wine"
    current["C1"] = "Tank"
    current["D1"] = "Sugar"
    current["E1"] = "pH"
    current["F1"] = "Volume"
    current["G1"] = "State"
    current["H1"] = "SO2"
    # Set b2 for avoiding errors
    # current["B2"] = " "

    # Create section for schedule
    sch = book.create_sheet("Schedule")
    sch = book['Schedule']
    # Create section for
    book.create_sheet("Todo")
    todo = book["Todo"]
    # Create section for library
    book.create_sheet("Lib")
    lib = book["Lib"]
    lib["A1"] = "Index"
    lib["B1"] = "Wine"
    lib["C1"] = "Vintage"
    lib["D1"] = "Quantity"
    lib["E1"] = "Notes"
    # Reset main page as current
    book.save("WineMakerData.xlsx")
    # ws1 = wb.create_sheet("Wine Data")

customtkinter.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("green")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("WineMaker")
        self.geometry(f"{1100}x{580}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        # create sidebar frame with widgets-----------------------------------------------------------------------------
        # Create Sidebar frame
        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(6, weight=1)
        # Create main label
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="WineMaker",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        # Create Home button
        self.home_btn = customtkinter.CTkButton(self.sidebar_frame, command=self.home_screen)
        self.home_btn.grid(row=1, column=0, padx=20, pady=10)
        self.home_btn.configure(text="Home")
        # Create Wines button
        self.wines_btn = customtkinter.CTkButton(self.sidebar_frame, command=self.wine_screen)
        self.wines_btn.grid(row=2, column=0, padx=20, pady=10)
        self.wines_btn.configure(text="Wines")
        # Create Calculations button
        self.calc_btn = customtkinter.CTkButton(self.sidebar_frame, command=self.calc_screen)
        self.calc_btn.grid(row=3, column=0, padx=20, pady=10)
        self.calc_btn.configure(text="Calculations")
        # Create Library button
        self.lib_btn = customtkinter.CTkButton(self.sidebar_frame, command=self.lib_screen)
        self.lib_btn.grid(row=4, column=0, padx=20, pady=10)
        self.lib_btn.configure(text="Library")
        # Create Schedule button
        self.sch_btn = customtkinter.CTkButton(self.sidebar_frame, command=self.schedule_page)
        self.sch_btn.grid(row=5, column=0, padx=20, pady=10)
        self.sch_btn.configure(text="Schedule")
        # Create Notes button
        # self.note_btn = customtkinter.CTkButton(self.sidebar_frame, command=self.sidebar_button_event)
        # self.note_btn.grid(row=5, column=0, padx=20, pady=10)
        # self.note_btn.configure(text="Notes")
        # Create apperance label
        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        # Create apperance dropdown
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame,
                                                                       values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 10))
        # ---------------------------------------------------------------------------------------------------------------------------
        # # Run loop to fill array with indexes, this is for indexing purposes
        # self.index_array = []
        # for i in range(self.get_maximum_rows(sheet_object=current) - 1):
        #     self.index_array.append(2+i)

        self.home_screen()

        # for i in range(3):
        #     current["A" + str(i+2)] = 1
        #     # print("A" + str(i+2))
        #
        # for i in range(self.get_maximum_rows(sheet_object=current)-1):
        #     current["B" + str(i + 2)] = 2
        #     current["C" + str(i + 2)] = 3
        #     current["D" + str(i + 2)] = 4
        #
        # book.save("WineMakerData.xlsx")

        # current["C9"] = 66

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def sidebar_button_event(self):
        print("sidebar_button click")

    def home_screen(self):
        book.active = book['WineData']
        book.save("WineMakerData.xlsx")
        self.main_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.main_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")
        # self.main_frame.grid_rowconfigure(6, weight=1)

        # Home label
        self.home_label = customtkinter.CTkLabel(self.main_frame, text="Home",
                                                 font=customtkinter.CTkFont(size=30, weight="bold"))
        self.home_label.grid(row=0, column=1, padx=(0, 500), pady=(20, 400))
        # self.sidebar_frame.grid_rowconfigure(2, weight=0)

        # To-do---------------------------------------------------------------------------------------
        # to-do label
        self.todo_label = customtkinter.CTkLabel(self.main_frame, text="To-Do",
                                                 font=customtkinter.CTkFont(size=30, weight="bold"))
        self.todo_label.place(x=0, y=70)
        self.add_todo = customtkinter.CTkButton(self.main_frame, text="+", width=20, height=20,
                                                command=self.create_todo)
        self.add_todo.place(x=100, y=80)
        # to-do list
        self.scrollable_todo = customtkinter.CTkScrollableFrame(self.main_frame, width=500, height=400)
        self.scrollable_todo.place(x=0, y=120)

        # # Set todos from schedule
        # today = date.today()
        # for i in range(self.get_maximum_rows(sheet_object=sch)):
        #     if(int(sch["A" + str(i + 1)].value[0]+sch["A" + str(i + 1)].value[1]+sch["A" + str(i + 1)].value[2]+sch["A" + str(i + 1)].value[3]) <= today.year):
        #         if(sch["A" + str(i + 1)].value[6] == '-'):
        #             if (int(sch["A" + str(i + 1)].value[5]) <= today.month):
        #                 if(len(sch["A" + str(i + 1)].value) >= 8):
        #                     if(int(sch["A" + str(i + 1)].value[7]) <= today.month):
        #                         todo["A" + str(self.get_maximum_rows(sheet_object=todo) + 1)] = sch["B" + str(i + 1)].value
        #                 elif(len(sch["A" + str(i + 1)].value) >= 9):
        #                     if (int(sch["A" + str(i + 1)].value[7] + sch["A" + str(i + 1)].value[8]) <= today.month):
        #                         todo["A" + str(self.get_maximum_rows(sheet_object=todo) + 1 )] = sch["B" + str(i + 1)].value
        #         else:
        #             if(len(sch["A" + str(i + 1)].value) >= 9):
        #                 if (int(sch["A" + str(i + 1)].value[8]) <= today.month):
        #                     todo["A" + str(self.get_maximum_rows(sheet_object=todo) + 1 )] = sch["B" + str(i + 1)].value
        #             elif(len(sch["A" + str(i + 1)].value) >= 10):
        #                 if (int(sch["A" + str(i + 1)].value[8] + sch["A" + str(i + 1)].value[9]) <= today.month):
        #                     todo["A" + str(self.get_maximum_rows(sheet_object=todo) + 1)] = sch["B" + str(i + 1)].value

        # Sets info for index purposes
        self.info_name = current["B2"].value

        self.todo_btns = []

        for i in range(self.get_maximum_rows(sheet_object=todo)):
            # wine
            todo_list = customtkinter.CTkLabel(master=self.scrollable_todo,
                                               text=str(i + 1) + ") " + str(todo["A" + str(i + 1)].value),
                                               font=customtkinter.CTkFont(size=20, weight="bold"))
            todo_list.grid(row=i, column=0, padx=0, pady=(0, 20))
            self.scrollable_todo.columnconfigure(1, weight=1)

            complete = customtkinter.CTkButton(self.scrollable_todo, text="Done" + str(i + 1), width=60, height=30,
                                               command=lambda c=i: self.delete_todo(c))
            complete.grid(row=i, column=2, padx=10, pady=(0, 20))

            self.todo_btns.append(complete)

        # Tank occupancy--------------------------------------------------------------------------------------
        # Tank label
        self.tank_home_label = customtkinter.CTkLabel(self.main_frame, text="Tanks:",
                                                      font=customtkinter.CTkFont(size=30, weight="bold"))
        self.tank_home_label.place(x=540, y=70)
        # Loop for tank numbers
        self.tanks = []
        self.volumes = []
        for i in range(11):
            # Main tank label
            home_tank_num = customtkinter.CTkLabel(self.main_frame, text="Tank " + str(i + 1) + ":",
                                                   font=customtkinter.CTkFont(size=20))
            home_tank_num.place(x=550, y=((120) + (i * 40)))
            # Wine name
            home_wine_name = customtkinter.CTkLabel(self.main_frame,
                                                    text="empty")  # , font=customtkinter.CTkFont(size=20))
            home_wine_name.place(x=670, y=((120) + (i * 40)))
            self.tanks.append(home_wine_name)
            # Volume at tank
            home_wine_vol = customtkinter.CTkLabel(self.main_frame,
                                                   text="0 L")  # , font=customtkinter.CTkFont(size=20))
            home_wine_vol.place(x=800, y=((120) + (i * 40)))
            self.volumes.append(home_wine_vol)

        for i in range(self.get_maximum_rows(sheet_object=current) - 1):
            if (current["C" + str(i + 2)].value < 12):
                self.tanks[current["C" + str(i + 2)].value - 1].configure(text=current["B" + str(i + 2)].value, font=customtkinter.CTkFont(size=20, weight="bold"))
                self.volumes[current["C" + str(i + 2)].value - 1].configure(
                    text=str(current["F" + str(i + 2)].value) + " L",
                    font=customtkinter.CTkFont(size=20, weight="bold"))

    def create_todo(self):
        self.todo_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.todo_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        self.add_todo_label = customtkinter.CTkLabel(self.todo_frame, text="New To-Do",
                                                     font=customtkinter.CTkFont(size=30, weight="bold"))
        self.add_todo_label.place(x=0, y=20)

        self.info_label = customtkinter.CTkLabel(self.todo_frame, text="To-do Information",
                                                 font=customtkinter.CTkFont(size=20))
        self.info_label.place(x=0, y=70)

        self.todo_item = customtkinter.CTkTextbox(self.todo_frame, width=500, height=200)
        self.todo_item.place(x=0, y=100)

        self.confirm_todo = customtkinter.CTkButton(self.todo_frame, text="Confirm", command=self.new_todo)
        self.confirm_todo.place(x=0, y=350)

    def new_todo(self):
        todo["A" + str(self.get_maximum_rows(sheet_object=todo) + 1)] = self.todo_item.get("0.0", "end")
        book.save("WineMakerData.xlsx")
        self.home_screen()

    def delete_todo(self, indx):
        todo.delete_rows(indx + 1)
        book.save("WineMakerData.xlsx")
        self.home_screen()

    def wine_screen(self):
        book.active = book['WineData']
        book.save("WineMakerData.xlsx")
        # Create main frame
        self.wine_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.wine_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        # Wine label
        self.wines_label = customtkinter.CTkLabel(self.wine_frame, text="Wines",
                                                  font=customtkinter.CTkFont(size=30, weight="bold"))
        self.wines_label.place(x=20, y=20)

        # Create wine tabs-------------------------------------------------------------------------
        # tabs should include
        # wine name
        # Tank number
        # Volume
        # Stage
        # ph
        # SO2

        # wine
        self.wine_label = customtkinter.CTkLabel(self.wine_frame, text="Wine",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.wine_label.place(x=40, y=100)
        # tank
        self.tank_label = customtkinter.CTkLabel(self.wine_frame, text="Tank(s)",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.tank_label.place(x=150, y=100)
        # volume
        self.vol_label = customtkinter.CTkLabel(self.wine_frame, text="Volume",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.vol_label.place(x=290, y=100)
        # stage
        self.stage_label = customtkinter.CTkLabel(self.wine_frame, text="Stage",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.stage_label.place(x=460, y=100)

        # pH
        self.ph_label = customtkinter.CTkLabel(self.wine_frame, text="pH",
                                               font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ph_label.place(x=600, y=100)
        # SO2
        self.so2_label = customtkinter.CTkLabel(self.wine_frame, text="SO2",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.so2_label.place(x=710, y=100)
        # Add button
        self.addWine = customtkinter.CTkButton(self.wine_frame, text="+", width=20, height=20, command=self.add_wine)
        self.addWine.place(x=845, y=110)
        # Bottom Line
        self.wine_line = customtkinter.CTkLabel(self.wine_frame,
                                                text="---------------------------------------------------------------------------------------------------------------------",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.wine_line.place(x=20, y=130)
        # ---------------------------------------------------------------------------------------------------------
        # Presets edit combo box value
        self.info_name = current["B2"].value

        # Add scrollable frame
        self.scrollable_frame = customtkinter.CTkScrollableFrame(self.wine_frame, width=850, height=400)
        self.scrollable_frame.place(x=20, y=150)

        # Loop for adding wines and values
        for i in range(self.get_maximum_rows(sheet_object=current) - 1):
            # wine
            wine = customtkinter.CTkLabel(master=self.scrollable_frame, text=current["B" + str(i + 2)].value,
                                          font=customtkinter.CTkFont(size=20, weight="bold"))
            wine.grid(row=i, column=0, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(1, weight=1)
            # tank
            tank = customtkinter.CTkLabel(master=self.scrollable_frame, text=current["C" + str(i + 2)].value,
                                          font=customtkinter.CTkFont(size=20, weight="bold"))
            tank.grid(row=i, column=2, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(3, weight=1)
            # volume
            volume = customtkinter.CTkLabel(master=self.scrollable_frame,
                                            text=str(current["F" + str(i + 2)].value) + " L",
                                            font=customtkinter.CTkFont(size=20, weight="bold"))
            volume.grid(row=i, column=4, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(5, weight=1)
            # Stage
            stage = customtkinter.CTkLabel(master=self.scrollable_frame,
                                           text=current["G" + str(i + 2)].value,
                                           font=customtkinter.CTkFont(size=20, weight="bold"))
            stage.grid(row=i, column=6, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(7, weight=1)
            # pH
            ph = customtkinter.CTkLabel(master=self.scrollable_frame, text=current["E" + str(i + 2)].value,
                                        font=customtkinter.CTkFont(size=20, weight="bold"))
            ph.grid(row=i, column=8, padx=8, pady=(0, 20))
            self.scrollable_frame.columnconfigure(9, weight=1)
            # SO2
            so2 = customtkinter.CTkLabel(master=self.scrollable_frame,
                                         text=str(current["H" + str(i + 2)].value) + " ppm",
                                         font=customtkinter.CTkFont(size=20, weight="bold"))
            so2.grid(row=i, column=10, padx=10, pady=(0, 20))

            # edit button
            self.scrollable_frame.columnconfigure(11, weight=1)
            edit = customtkinter.CTkButton(self.scrollable_frame, text="...", width=20, height=15,
                                           command=lambda: self.wine_info())  # (int(current["A" + str(i+2)].value) + 1))
            edit.grid(row=i, column=12, padx=10, pady=(0, 20))

    def add_wine(self):
        # Create main frame
        self.add_wine_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.add_wine_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        # Wine label
        self.add_wines_label = customtkinter.CTkLabel(self.add_wine_frame, text="New Wines",
                                                      font=customtkinter.CTkFont(size=30, weight="bold"))
        self.add_wines_label.place(x=20, y=20)

        # Input Wine name
        self.new_name = customtkinter.CTkLabel(self.add_wine_frame, text="Name:",
                                               font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_name.place(x=20, y=100)

        self.name_entry = customtkinter.CTkEntry(self.add_wine_frame, placeholder_text="Name")
        self.name_entry.place(x=120, y=100)
        self.name_entry.insert(0, "New")

        # Input tank number
        self.new_tank = customtkinter.CTkLabel(self.add_wine_frame, text="Tank:",
                                               font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_tank.place(x=20, y=150)

        self.tank_entry = customtkinter.CTkEntry(self.add_wine_frame, placeholder_text="Tank Number")
        self.tank_entry.place(x=120, y=150)
        self.tank_entry.insert(0, "0")

        # Input Volume
        self.new_vol = customtkinter.CTkLabel(self.add_wine_frame, text="Volume:",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_vol.place(x=20, y=200)

        self.vol_entry = customtkinter.CTkEntry(self.add_wine_frame, placeholder_text="Volume")
        self.vol_entry.place(x=120, y=200)
        self.vol_entry.insert(0, "0")

        self.vol_l_label = customtkinter.CTkLabel(self.add_wine_frame, text=" L",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.vol_l_label.place(x=270, y=200)

        # Input Sugar
        self.new_sugar = customtkinter.CTkLabel(self.add_wine_frame, text="Sugar:",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_sugar.place(x=20, y=250)

        self.sugar_entry = customtkinter.CTkEntry(self.add_wine_frame, placeholder_text="Sugar")
        self.sugar_entry.place(x=120, y=250)
        self.sugar_entry.insert(0, "0")

        self.sugar_label = customtkinter.CTkLabel(self.add_wine_frame, text=" g/L",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.sugar_label.place(x=270, y=250)

        # Input Wine pH
        self.new_ph = customtkinter.CTkLabel(self.add_wine_frame, text="pH:",
                                             font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_ph.place(x=20, y=300)

        self.ph_entry = customtkinter.CTkEntry(self.add_wine_frame, placeholder_text="pH")
        self.ph_entry.place(x=120, y=300)
        self.ph_entry.insert(0, "0")
        # Input Wine state
        self.new_state = customtkinter.CTkLabel(self.add_wine_frame, text="Stage:",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_state.place(x=20, y=350)

        self.state_entry = customtkinter.CTkEntry(self.add_wine_frame, placeholder_text="State")
        self.state_entry.place(x=120, y=350)
        self.state_entry.insert(0, "New")
        # Input Wine SO2
        self.new_so2 = customtkinter.CTkLabel(self.add_wine_frame, text="SO2:",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_so2.place(x=20, y=400)

        self.so2_entry = customtkinter.CTkEntry(self.add_wine_frame, placeholder_text="SO2")
        self.so2_entry.place(x=120, y=400)
        self.so2_entry.insert(0, "0")

        self.so2_ppm_label = customtkinter.CTkLabel(self.add_wine_frame, text=" ppm",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.so2_ppm_label.place(x=270, y=400)
        # Confirm Button
        self.confirm_btn = customtkinter.CTkButton(self.add_wine_frame, text="Confirm", command=self.confirm_add)
        self.confirm_btn.place(x=20, y=450)

    def wine_info(self):
        # Create main frame
        self.wine_info_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.wine_info_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        # Info label
        self.wines_info_label = customtkinter.CTkLabel(self.wine_info_frame, text="Wine Information",
                                                       font=customtkinter.CTkFont(size=30, weight="bold"))
        self.wines_info_label.place(x=20, y=20)
        # Wine dropdown bar
        self.wine_drop = customtkinter.CTkComboBox(self.wine_info_frame,
                                                   values=[current["B" + str(i + 2)].value for i in
                                                           range(self.get_maximum_rows(sheet_object=current) - 1)],
                                                   command=self.set_info)
        self.wine_drop.place(x=320, y=25)
        self.wine_drop.set(self.info_name)
        for i in range(self.get_maximum_rows(sheet_object=current) - 1):  # in progress loop for combo box index
            if current["B" + str(i + 2)].value == self.wine_drop.get():
                self.info_index = str(i + 2)

        # Edit button
        self.edit_btn = customtkinter.CTkButton(self.wine_info_frame, text="Edit", command=self.edit_wine)
        self.edit_btn.place(x=600, y=25)

        # Show Wine name
        self.wine_name = customtkinter.CTkLabel(self.wine_info_frame, text="Name:",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.wine_name.place(x=20, y=100)

        self.current_wine = customtkinter.CTkLabel(self.wine_info_frame, text=current["B" + self.info_index].value)
        self.current_wine.place(x=120, y=100)

        # Show tank number
        self.tank_num = customtkinter.CTkLabel(self.wine_info_frame, text="Tank:",
                                               font=customtkinter.CTkFont(size=20, weight="bold"))
        self.tank_num.place(x=20, y=150)

        self.current_tank = customtkinter.CTkLabel(self.wine_info_frame, text=current["C" + self.info_index].value)
        self.current_tank.place(x=120, y=150)

        # Show Volume
        self.vol_num = customtkinter.CTkLabel(self.wine_info_frame, text="Volume:",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
        self.vol_num.place(x=20, y=200)

        self.current_vol = customtkinter.CTkLabel(self.wine_info_frame,
                                                  text=str(current["F" + self.info_index].value) + " L")
        self.current_vol.place(x=120, y=200)

        # Show Sugar
        self.sugar_num = customtkinter.CTkLabel(self.wine_info_frame, text="Sugar:",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.sugar_num.place(x=20, y=250)

        self.current_sugar = customtkinter.CTkLabel(self.wine_info_frame,
                                                    text=str(current["D" + self.info_index].value) + " g/L")
        self.current_sugar.place(x=120, y=250)

        # Show Wine pH
        self.ph_num = customtkinter.CTkLabel(self.wine_info_frame, text="pH:",
                                             font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ph_num.place(x=20, y=300)

        self.current_ph = customtkinter.CTkLabel(self.wine_info_frame, text=current["E" + self.info_index].value)
        self.current_ph.place(x=120, y=300)
        # Show Wine state
        self.state_name = customtkinter.CTkLabel(self.wine_info_frame, text="Stage:",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.state_name.place(x=20, y=350)

        self.current_state = customtkinter.CTkLabel(self.wine_info_frame, text=current["G" + self.info_index].value)
        self.current_state.place(x=120, y=350)
        # Show Wine so2
        self.so2_num = customtkinter.CTkLabel(self.wine_info_frame, text="SO2:",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
        self.so2_num.place(x=20, y=400)

        self.current_so2 = customtkinter.CTkLabel(self.wine_info_frame,
                                                  text=str(current["H" + self.info_index].value) + " ppm")
        self.current_so2.place(x=120, y=400)

    def edit_wine(self):
        # Create main frame
        self.edit_wine_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.edit_wine_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        # Wine label
        self.edit_wines_label = customtkinter.CTkLabel(self.edit_wine_frame, text="Edit Wine",
                                                       font=customtkinter.CTkFont(size=30, weight="bold"))
        self.edit_wines_label.place(x=20, y=20)

        # Delete button
        self.delete_btn = customtkinter.CTkButton(self.edit_wine_frame, text="Delete wine", command=self.you_sure,
                                                  fg_color="red", hover_color="red")
        self.delete_btn.place(x=700, y=25)

        # Input Wine name
        self.edit_name = customtkinter.CTkLabel(self.edit_wine_frame, text="Name:",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_name.place(x=20, y=100)

        self.edit_name_entry = customtkinter.CTkEntry(self.edit_wine_frame, placeholder_text="Name")
        self.edit_name_entry.place(x=120, y=100)
        self.edit_name_entry.insert(0, current["B" + self.info_index].value)

        # Input tank number
        self.edit_new_tank = customtkinter.CTkLabel(self.edit_wine_frame, text="Tank:",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_new_tank.place(x=20, y=150)

        self.edit_tank_entry = customtkinter.CTkEntry(self.edit_wine_frame, placeholder_text="Tank Number")
        self.edit_tank_entry.place(x=120, y=150)
        self.edit_tank_entry.insert(0, current["C" + self.info_index].value)

        # Input Volume
        self.edit_new_vol = customtkinter.CTkLabel(self.edit_wine_frame, text="Volume:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_new_vol.place(x=20, y=200)

        self.edit_vol_entry = customtkinter.CTkEntry(self.edit_wine_frame, placeholder_text="Volume")
        self.edit_vol_entry.place(x=120, y=200)
        self.edit_vol_entry.insert(0, current["F" + self.info_index].value)

        self.edit_l_label = customtkinter.CTkLabel(self.edit_wine_frame, text=" L",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_l_label.place(x=270, y=200)

        # Input Sugar
        self.edit_new_sugar = customtkinter.CTkLabel(self.edit_wine_frame, text="Sugar:",
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_new_sugar.place(x=20, y=250)

        self.edit_sugar_entry = customtkinter.CTkEntry(self.edit_wine_frame, placeholder_text="Sugar")
        self.edit_sugar_entry.place(x=120, y=250)
        self.edit_sugar_entry.insert(0, current["D" + self.info_index].value)

        self.edit_g_label = customtkinter.CTkLabel(self.edit_wine_frame, text=" g/L",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_g_label.place(x=270, y=250)

        # Input Wine pH
        self.edit_new_ph = customtkinter.CTkLabel(self.edit_wine_frame, text="pH:",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_new_ph.place(x=20, y=300)

        self.edit_ph_entry = customtkinter.CTkEntry(self.edit_wine_frame, placeholder_text="pH")
        self.edit_ph_entry.place(x=120, y=300)
        self.edit_ph_entry.insert(0, current["E" + self.info_index].value)
        # Input Wine state
        self.edit_new_state = customtkinter.CTkLabel(self.edit_wine_frame, text="Stage:",
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_new_state.place(x=20, y=350)

        self.edit_state_entry = customtkinter.CTkEntry(self.edit_wine_frame, placeholder_text="State")
        self.edit_state_entry.place(x=120, y=350)
        self.edit_state_entry.insert(0, current["G" + self.info_index].value)
        # Input Wine SO2
        self.edit_new_so2 = customtkinter.CTkLabel(self.edit_wine_frame, text="SO2:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_new_so2.place(x=20, y=400)

        self.edit_so2_entry = customtkinter.CTkEntry(self.edit_wine_frame, placeholder_text="SO2")
        self.edit_so2_entry.place(x=120, y=400)
        self.edit_so2_entry.insert(0, current["H" + self.info_index].value)

        self.edit_so2_label = customtkinter.CTkLabel(self.edit_wine_frame, text=" ppm",
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
        self.edit_so2_label.place(x=270, y=400)
        # Confirm Button
        self.edit_confirm_btn = customtkinter.CTkButton(self.edit_wine_frame, text="Confirm", command=self.confirm_edit)
        self.edit_confirm_btn.place(x=20, y=450)

    def set_info(self, val: str):
        self.info_name = self.wine_drop.get()
        self.wine_info()

    def confirm_add(self):
        ind = self.get_maximum_rows(sheet_object=current) + 1

        try:
            int(self.vol_entry.get())
        except ValueError:
            # Error message
            self.error_vol = customtkinter.CTkLabel(self.add_wine_frame, text="Error, must input a number")
            self.error_vol.place(x=350, y=200)
            return

        try:
            int(self.tank_entry.get())
        except ValueError:
            # Error message
            self.error_tank = customtkinter.CTkLabel(self.add_wine_frame, text="Error, must input a number")
            self.error_tank.place(x=350, y=150)

        try:
            self.name_entry.get()
        except ValueError:
            # Error message
            self.error_name = customtkinter.CTkLabel(self.add_wine_frame, text="Error, must input a name")
            self.error_name.place(x=350, y=100)
            return

        try:
            int(self.sugar_entry.get())
        except ValueError:
            # Error message
            self.error_sugar = customtkinter.CTkLabel(self.add_wine_frame, text="Error, must input a number")
            self.error_sugar.place(x=350, y=250)
            return

        try:
            int(self.ph_entry.get())
        except ValueError:
            # Error message
            self.error_ph = customtkinter.CTkLabel(self.add_wine_frame, text="Error, must input a number")
            self.error_ph.place(x=350, y=300)
            return

        try:
            self.state_entry.get()
        except ValueError:
            # Error message
            self.error_state = customtkinter.CTkLabel(self.add_wine_frame, text="Error, must input a state")
            self.error_state.place(x=350, y=350)
            return

        try:
            int(self.so2_entry.get())
        except ValueError:
            # Error message
            self.error_so2 = customtkinter.CTkLabel(self.add_wine_frame, text="Error, must input a number")
            self.error_so2.place(x=350, y=400)
            return

        current["A" + str(ind)] = ind - 1
        current["B" + str(ind)] = self.name_entry.get()
        current["C" + str(ind)] = int(self.tank_entry.get())
        current["D" + str(ind)] = int(self.sugar_entry.get())
        current["E" + str(ind)] = float(self.ph_entry.get())
        current["F" + str(ind)] = int(self.vol_entry.get())
        current["G" + str(ind)] = self.state_entry.get()
        current["H" + str(ind)] = float(self.so2_entry.get())
        book.save("WineMakerData.xlsx")
        self.wine_screen()

    def confirm_edit(self):
        ind = int(self.info_index)

        try:
            int(self.edit_vol_entry.get())
        except ValueError:
            # Error message
            self.error_vol = customtkinter.CTkLabel(self.edit_wine_frame, text="Error, must input a number")
            self.error_vol.place(x=350, y=200)
            return

        try:
            int(self.edit_tank_entry.get())
        except ValueError:
            # Error message
            self.error_tank = customtkinter.CTkLabel(self.edit_wine_frame, text="Error, must input a number")
            self.error_tank.place(x=350, y=150)
            return

        current["A" + str(ind)] = ind - 1
        current["B" + str(ind)] = self.edit_name_entry.get()
        current["C" + str(ind)] = int(self.edit_tank_entry.get())
        current["D" + str(ind)] = int(self.edit_sugar_entry.get())
        current["E" + str(ind)] = float(self.edit_ph_entry.get())
        current["F" + str(ind)] = int(self.edit_vol_entry.get())
        current["G" + str(ind)] = self.edit_state_entry.get()
        current["H" + str(ind)] = float(self.edit_so2_entry.get())
        book.save("WineMakerData.xlsx")
        self.wine_screen()

    def delete_wine(self):
        current.delete_rows(int(self.info_index))
        book.save("WineMakerData.xlsx")
        self.wine_screen()

    def you_sure(self):
        self.delete1_btn = customtkinter.CTkButton(self.edit_wine_frame, text="Are you sure?",
                                                   command=self.you_really_sure,
                                                   fg_color="red", hover_color="red")
        self.delete1_btn.place(x=700, y=500)

    def you_really_sure(self):
        self.delete2_btn = customtkinter.CTkButton(self.edit_wine_frame, text="Final chance", command=self.delete_wine,
                                                   fg_color="red", hover_color="red")
        self.delete2_btn.place(x=700, y=250)

    def calc_screen(self):
        book.active = book['WineData']
        book.save("WineMakerData.xlsx")
        # Create main frame
        self.calc_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.calc_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        # Create tabs
        self.calctabs = customtkinter.CTkTabview(self.calc_frame, width=850, height=500)
        self.calctabs.place(x=20, y=50)
        self.calctabs.add("SO2")
        self.calctabs.add("Fermentation")
        self.calctabs.add("Conversions")

        # Calc label
        self.calc_label = customtkinter.CTkLabel(self.calc_frame, text="Calculations",
                                                 font=customtkinter.CTkFont(size=30, weight="bold"))
        self.calc_label.place(x=20, y=20)

        # Wine dropdown bar
        self.wine_drop = customtkinter.CTkComboBox(self.calc_frame, values=[current["B" + str(i + 2)].value for i in
                                                                            range(self.get_maximum_rows(
                                                                                sheet_object=current) - 1)],
                                                   command=self.set_calc_info)
        self.wine_drop.place(x=320, y=25)
        try:
            self.wine_drop.set(self.info_name)
        except tkinter.TclError:
            pass

        for i in range(self.get_maximum_rows(sheet_object=current) - 1):  # in progress loop for combo box index
            if current["B" + str(i + 2)].value == self.wine_drop.get():
                info_index = str(i + 2)

        # SO2 Calculation---------------------------------------------------------------------------------------------
        # Show Wine name
        self.s02_wine_name = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="Name:",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.s02_wine_name.place(x=20, y=0)

        self.s02_current_wine = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text=current["B" + info_index].value)
        self.s02_current_wine.place(x=200, y=0)

        # Show tank number
        self.s02_tank_num = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="Tank:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.s02_tank_num.place(x=20, y=50)

        self.s02_current_tank = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text=current["C" + info_index].value)
        self.s02_current_tank.place(x=200, y=50)

        # Show Volume
        self.s02_vol_num = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="Volume:",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.s02_vol_num.place(x=20, y=100)

        self.s02_current_vol = customtkinter.CTkEntry(
            self.calctabs.tab("SO2"))  # , text=current["F" + info_index].value)
        self.s02_current_vol.place(x=200, y=100)
        self.s02_current_vol.insert(0, current["F" + info_index].value)
        self.l_label = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="L")
        self.l_label.place(x=350, y=100)
        # Add SO2 addition
        self.s02_s02_num = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="SO2 Addition:",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.s02_s02_num.place(x=20, y=150)
        self.s02_added_s02 = customtkinter.CTkEntry(self.calctabs.tab("SO2"))  # , text=current["F" + info_index].value)
        self.s02_added_s02.place(x=200, y=150)
        self.ppm_label = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="ppm")
        self.ppm_label.place(x=350, y=150)
        # SO2 result
        self.s02_result = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="Meta Addition:",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.s02_result.place(x=20, y=200)

        self.s02_current_result = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="answer")
        self.s02_current_result.place(x=200, y=200)

        self.sO2_calculate = customtkinter.CTkButton(self.calctabs.tab("SO2"), text="Calculate", command=self.calc_so2)
        self.sO2_calculate.place(x=20, y=250)

        # Fermentation Calculation-----------------------------------------------------------------------------------
        # Show Wine name
        self.ferm_wine_name = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Name:",
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ferm_wine_name.place(x=20, y=0)

        self.ferm_current_wine = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"),
                                                        text=current["B" + info_index].value)
        self.ferm_current_wine.place(x=200, y=0)

        # Show tank number
        self.ferm_tank_num = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Tank:",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ferm_tank_num.place(x=20, y=50)

        self.ferm_current_tank = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"),
                                                        text=current["C" + info_index].value)
        self.ferm_current_tank.place(x=200, y=50)

        # Show Volume
        self.ferm_vol_num = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Volume:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ferm_vol_num.place(x=20, y=100)

        self.ferm_current_vol = customtkinter.CTkEntry(
            self.calctabs.tab("Fermentation"))  # , text=current["F" + info_index].value)
        self.ferm_current_vol.place(x=200, y=100)
        self.ferm_current_vol.insert(0, current["F" + info_index].value)
        self.l_label = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="L")
        self.l_label.place(x=350, y=100)
        # Add Brix
        self.ferm_brix_num = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Brix:",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ferm_brix_num.place(x=20, y=150)
        self.ferm_added_brix = customtkinter.CTkEntry(
            self.calctabs.tab("Fermentation"))  # , text=current["F" + info_index].value)
        self.ferm_added_brix.place(x=200, y=150)
        # Add Desired alcohol
        self.ferm_alc_num = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Desired Alcohol:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ferm_alc_num.place(x=20, y=200)
        self.ferm_des_alc = customtkinter.CTkEntry(
            self.calctabs.tab("Fermentation"))  # , text=current["F" + info_index].value)
        self.ferm_des_alc.place(x=200, y=200)
        # Add yeast per hl
        self.ferm_alc_num = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Yeast g/hL:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ferm_alc_num.place(x=20, y=250)
        self.yeast_drop = customtkinter.CTkComboBox(self.calctabs.tab("Fermentation"), values=["25", "35"])
        self.yeast_drop.place(x=200, y=250)
        # Calculate button
        self.ferm_calculate = customtkinter.CTkButton(self.calctabs.tab("Fermentation"), text="Calculate",
                                                      command=self.calc_ferm)
        self.ferm_calculate.place(x=20, y=300)
        # Current Alcohol result
        self.alc_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Current Alcohol:",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.alc_result.place(x=450, y=0)

        self.alc_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.alc_current_result.place(x=630, y=0)
        # Volume gallon result
        self.vol_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Volume(gal):",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.vol_result.place(x=450, y=40)

        self.vol_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.vol_current_result.place(x=630, y=40)
        # Yeast result
        self.yeast_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Yeast:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.yeast_result.place(x=450, y=80)

        self.yeast_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.yeast_current_result.place(x=630, y=80)
        # Water result
        self.water_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Water needed:",
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.water_result.place(x=450, y=120)

        self.water_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.water_current_result.place(x=630, y=120)
        # Fermaid result
        self.fermaid_lb_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Fermaid(lbs):",
                                                        font=customtkinter.CTkFont(size=20, weight="bold"))
        self.fermaid_lb_result.place(x=450, y=160)

        self.fermaid_lb_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.fermaid_lb_current_result.place(x=630, y=160)
        self.fermaid_g_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Fermaid(g):",
                                                       font=customtkinter.CTkFont(size=20, weight="bold"))
        self.fermaid_g_result.place(x=450, y=200)

        self.fermaid_g_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.fermaid_g_current_result.place(x=630, y=200)
        # Yeast Hulls result
        self.yeasthull_lb_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Yeast Hulls(lbs):",
                                                          font=customtkinter.CTkFont(size=20, weight="bold"))
        self.yeasthull_lb_result.place(x=450, y=240)

        self.yeasthull_lb_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.yeasthull_lb_current_result.place(x=630, y=240)
        self.yeasthull_g_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Yeast Hulls(g):",
                                                         font=customtkinter.CTkFont(size=20, weight="bold"))
        self.yeasthull_g_result.place(x=450, y=280)

        self.yeasthull_g_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.yeasthull_g_current_result.place(x=630, y=280)
        # Sugar result
        self.sugar_lb_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Sugar(lbs):",
                                                      font=customtkinter.CTkFont(size=20, weight="bold"))
        self.sugar_lb_result.place(x=450, y=320)

        self.sugar_lb_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.sugar_lb_current_result.place(x=630, y=320)
        self.sugar_g_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text="Sugar(Kg):",
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
        self.sugar_g_result.place(x=450, y=360)

        self.sugar_g_current_result = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.sugar_g_current_result.place(x=630, y=360)

        # Standard conversion Calculation-----------------------------------------------------------------------------------
        # Show Wine name
        self.con_vol = customtkinter.CTkLabel(self.calctabs.tab("Conversions"), text="Volume:",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
        self.con_vol.place(x=20, y=0)
        self.con_current_vol = customtkinter.CTkEntry(
            self.calctabs.tab("Conversions"))  # , text=current["F" + info_index].value)
        self.con_current_vol.place(x=200, y=0)
        self.con_current_vol.insert(0, current["F" + info_index].value)

        self.l_label = customtkinter.CTkLabel(self.calctabs.tab("Conversions"), text="L")
        self.l_label.place(x=350, y=0)

        self.con_drop = customtkinter.CTkComboBox(self.calctabs.tab("Conversions"), values=["g/hL", "g/L"])
        self.con_drop.place(x=380, y=50)

        self.con_den = customtkinter.CTkLabel(self.calctabs.tab("Conversions"), text="Density:",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
        self.con_den.place(x=20, y=50)
        self.con_current_den = customtkinter.CTkEntry(
            self.calctabs.tab("Conversions"))  # , text=current["F" + info_index].value)
        self.con_current_den.place(x=200, y=50)

        self.con_ans_label = customtkinter.CTkLabel(self.calctabs.tab("Conversions"), text="Grams:",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.con_ans_label.place(x=20, y=100)
        self.con_ans = customtkinter.CTkLabel(self.calctabs.tab("Conversions"), text=" ")
        self.con_ans.place(x=200, y=100)

        self.con_calculate = customtkinter.CTkButton(self.calctabs.tab("Conversions"), text="Calculate",
                                                     command=self.calc_con)
        self.con_calculate.place(x=20, y=150)

        # # Multiplication test
        # self.entry1 = customtkinter.CTkEntry(self.calctabs.tab("SO2"), placeholder_text="Num 1")
        # self.entry1.place(x=10, y=0)
        # # self.entry1.insert(0, current["C2"].value)
        # self.entry2 = customtkinter.CTkEntry(self.calctabs.tab("SO2"), placeholder_text="Num2")
        # self.entry2.place(x=200, y=0)
        # # self.entry2.insert(0, current["C3"].value)
        # # Mult label
        # self.mult_label = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="x", font=customtkinter.CTkFont(size=20, weight="bold"))
        # self.mult_label.place(x=170, y=0)
        # # equals label
        # self.eq_label = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="=",
        #                                          font=customtkinter.CTkFont(size=20, weight="bold"))
        # self.eq_label.place(x=360, y=0)
        #
        # # answer
        # self.an_label = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text="answer",
        #                                        font=customtkinter.CTkFont(size=20, weight="bold"))
        # self.an_label.place(x=390, y=0)
        #
        # # calculate button
        # self.get_an_btn = customtkinter.CTkButton(self.calctabs.tab("SO2"), width=80, command=self.mult)
        # self.get_an_btn.place(x=490, y=0)
        # self.get_an_btn.configure(text="Calculate")

    def set_calc_info(self, val: str):
        self.info_name = self.wine_drop.get()
        self.calc_screen()

    def calc_so2(self):
        self.so2_error_vol = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text=" ")
        self.so2_error_so2 = customtkinter.CTkLabel(self.calctabs.tab("SO2"), text=" ")
        self.so2_error_so2.place(x=380, y=150)
        self.so2_error_vol.place(x=380, y=100)
        try:
            float(self.s02_current_vol.get())
        except ValueError:
            # Error message
            self.so2_error_vol.configure(text="Error, must input a number")
            return

        try:
            float(self.s02_added_s02.get())
        except ValueError:
            # Error message
            self.so2_error_so2.configure(text="Error, must input a number")
            return

        self.s02_current_result.configure(text=str(self.round_half_up(
            ((float(self.s02_current_vol.get()) / 100)) * ((float(self.s02_added_s02.get()) / 10) * 2))) + " g")
        self.so2_error_so2.configure(text="                                                                   ")
        self.so2_error_vol.configure(text="                                                                   ")

    def calc_ferm(self):

        self.ferm_error_vol = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.ferm_error_brix = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.ferm_error_alc = customtkinter.CTkLabel(self.calctabs.tab("Fermentation"), text=" ")
        self.ferm_error_vol.place(x=20, y=400)
        self.ferm_error_brix.place(x=20, y=400)
        self.ferm_error_alc.place(x=20, y=400)
        try:
            float(self.ferm_current_vol.get())
        except ValueError:
            # Error message
            self.ferm_error_vol.configure(text="Error, must input a number in Volume")
            return

        try:
            float(self.ferm_added_brix.get())
        except ValueError:
            # Error message
            self.ferm_error_brix.configure(text="Error, must input a number in Brix")
            return

        try:
            float(self.ferm_des_alc.get())
        except ValueError:
            # Error message
            self.ferm_error_alc.configure(text="Error, must input a number in Alcohol")
            return

        cur_alc = float(self.ferm_added_brix.get()) * .58
        self.alc_current_result.configure(text=str(self.round_half_up(cur_alc, 4)) + " %")
        vol_gal = float(self.ferm_current_vol.get()) * .26417205
        self.vol_current_result.configure(text=str(self.round_half_up(vol_gal, 2)) + "  Gallons")
        added_yeast = (float(self.ferm_current_vol.get()) / 100) * float(self.yeast_drop.get())
        self.yeast_current_result.configure(text=str(self.round_half_up(added_yeast, 4)) + " g")
        self.water_current_result.configure(text=str(self.round_half_up((added_yeast / 100), 4)) + " L")
        fermaid_lb = (vol_gal / 1000) * 2
        self.fermaid_lb_current_result.configure(text=str(self.round_half_up(fermaid_lb, 2)) + " lbs")
        self.fermaid_g_current_result.configure(text=str(self.round_half_up((fermaid_lb * 453.59237), 2)) + " g")
        yeasthulls_lb = (vol_gal / 1000) * 3
        self.yeasthull_lb_current_result.configure(text=str(self.round_half_up(yeasthulls_lb, 2)) + " lbs")
        self.yeasthull_g_current_result.configure(text=str(self.round_half_up((yeasthulls_lb * 453.59237), 2)) + " g")
        sugar_kg = ((float(self.ferm_des_alc.get()) - cur_alc) * 1.8) * (float(self.ferm_current_vol.get()) / 100)
        self.sugar_lb_current_result.configure(
            text=str(self.round_half_up(((sugar_kg / 453.59237) * 1000), 4)) + " lbs")
        self.sugar_g_current_result.configure(text=str(self.round_half_up((sugar_kg), 4)) + " Kg")
        self.ferm_error_vol.configure(
            text="                                                                           ")

    def calc_con(self):

        self.con_error_vol = customtkinter.CTkLabel(self.calctabs.tab("Conversions"), text=" ")
        self.con_error_den = customtkinter.CTkLabel(self.calctabs.tab("Conversions"), text=" ")
        self.con_error_vol.place(x=20, y=300)
        self.con_error_den.place(x=20, y=300)
        try:
            float(self.con_current_vol.get())
        except ValueError:
            # Error message
            self.con_error_vol.configure(text="Error, must input a number")
            return

        try:
            float(self.con_current_den.get())
        except ValueError:
            # Error message
            self.con_error_den.configure(text="Error, must input a number")
            return

        if (self.con_drop.get() == "g/hL"):
            self.con_ans.configure(text=str(self.round_half_up(
                (float(self.con_current_vol.get()) / 100) * float(self.con_current_den.get()))) + " g")
        elif (self.con_drop.get() == "g/L"):
            self.con_ans.configure(text=str(
                self.round_half_up(float(self.con_current_vol.get()) * float(self.con_current_den.get()))) + " g")

        self.con_error_den.configure(text="                                                              ")

    # def mult(self):
    #     self.an_label.configure(text=(str(int(self.entry1.get()) * int(self.entry2.get()))))

    def lib_screen(self):
        book.active = book['Lib']
        book.save("WineMakerData.xlsx")
        # Create main frame
        self.lib_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.lib_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        # Wine label
        self.lib_wines_label = customtkinter.CTkLabel(self.lib_frame, text="Wine Library",
                                                      font=customtkinter.CTkFont(size=30, weight="bold"))
        self.lib_wines_label.place(x=20, y=20)

        # Create wine library tabs-------------------------------------------------------------------------
        # tabs should include
        # wine name
        # vintage
        # Quantity
        # Notes

        # wine
        self.lib_wine_label = customtkinter.CTkLabel(self.lib_frame, text="Wine",
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
        self.lib_wine_label.place(x=40, y=100)
        # Vinatage
        self.lib_vin_label = customtkinter.CTkLabel(self.lib_frame, text="Vintage",
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
        self.lib_vin_label.place(x=190, y=100)
        # Quantity
        self.lib_qt_label = customtkinter.CTkLabel(self.lib_frame, text="Quantity",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.lib_qt_label.place(x=400, y=100)
        # Notes
        self.lib_stage_label = customtkinter.CTkLabel(self.lib_frame, text="Notes",
                                                      font=customtkinter.CTkFont(size=20, weight="bold"))
        self.lib_stage_label.place(x=660, y=100)

        # Add button
        self.lib_addWine = customtkinter.CTkButton(self.lib_frame, text="+", width=20, height=20, command=self.add_wine_lib)
        self.lib_addWine.place(x=845, y=110)
        # Bottom Line
        self.lib_wine_line = customtkinter.CTkLabel(self.lib_frame,
                                                    text="---------------------------------------------------------------------------------------------------------------------",
                                                    font=customtkinter.CTkFont(size=20, weight="bold"))
        self.lib_wine_line.place(x=20, y=130)
        # ---------------------------------------------------------------------------------------------------------
        # Presets edit combo box value
        self.lib_info_name = current["B2"].value

        # Add scrollable frame
        self.lib_scrollable_frame = customtkinter.CTkScrollableFrame(self.lib_frame, width=850, height=400)
        self.lib_scrollable_frame.place(x=20, y=150)

        """ # Loop for adding wines and values
        for i in range(self.get_maximum_rows(sheet_object=current)-1):
            # wine
            wine = customtkinter.CTkLabel(master=self.scrollable_frame, text=current["B" + str(i+2)].value,
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
            wine.grid(row=i, column=0, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(1, weight=1)
            # tank
            tank = customtkinter.CTkLabel(master=self.scrollable_frame, text=current["C" + str(i+2)].value,
                                                     font=customtkinter.CTkFont(size=20, weight="bold"))
            tank.grid(row=i, column=2, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(3, weight=1)
            # volume
            volume = customtkinter.CTkLabel(master=self.scrollable_frame, text=str(current["F" + str(i+2)].value) + " L",
                                                      font=customtkinter.CTkFont(size=20, weight="bold"))
            volume.grid(row=i, column=4, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(5, weight=1)
            # Stage
            stage = customtkinter.CTkLabel(master=self.scrollable_frame,
                                            text=current["G" + str(i + 2)].value,
                                            font=customtkinter.CTkFont(size=20, weight="bold"))
            stage.grid(row=i, column=6, padx=10, pady=(0, 20))
            self.scrollable_frame.columnconfigure(7, weight=1)
            # pH
            ph = customtkinter.CTkLabel(master=self.scrollable_frame, text=current["E" + str(i+2)].value,
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
            ph.grid(row=i, column=8, padx=8, pady=(0, 20))
            self.scrollable_frame.columnconfigure(9, weight=1)
            # SO2
            so2 = customtkinter.CTkLabel(master=self.scrollable_frame, text=str(current["H" + str(i + 2)].value) + " ppm", font=customtkinter.CTkFont(size=20, weight="bold"))
            so2.grid(row=i, column=10, padx=10, pady=(0, 20))

            # edit button
            self.scrollable_frame.columnconfigure(11, weight=1)
            edit = customtkinter.CTkButton(self.scrollable_frame, text="...", width=20, height=15, command=lambda: self.wine_info())  # (int(current["A" + str(i+2)].value) + 1))
            edit.grid(row=i, column=12, padx=10, pady=(0, 20)) """

    def add_wine_lib(self):
        # Create main frame
        self.add_wine_lib_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.add_wine_lib_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")

        # Wine label
        self.add_wines_lib_label = customtkinter.CTkLabel(self.add_wine_lib_frame, text="New Wines",
                                                      font=customtkinter.CTkFont(size=30, weight="bold"))
        self.add_wines_lib_label.place(x=20, y=20)

        # Input Wine name
        self.new_name_lib = customtkinter.CTkLabel(self.add_wine_lib_frame, text="Name:",
                                               font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_name_lib.place(x=20, y=100)

        self.name_entry_lib = customtkinter.CTkEntry(self.add_wine_lib_frame, placeholder_text="Name")
        self.name_entry_lib.place(x=120, y=100)
        self.name_entry_lib.insert(0, "New")

        # Input Vintage
        self.new_vin_lib = customtkinter.CTkLabel(self.add_wine_lib_frame, text="Vintage:",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_vin_lib.place(x=20, y=150)

        self.vin_entry_lib = customtkinter.CTkEntry(self.add_wine_lib_frame, placeholder_text="Vintage")
        self.vin_entry_lib.place(x=120, y=150)
        self.vin_entry_lib.insert(0, "0")

        # Input Quantity
        self.new_qt_lib = customtkinter.CTkLabel(self.add_wine_lib_frame, text="Quantity:",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_qt_lib.place(x=20, y=200)

        self.qt_entry_lib = customtkinter.CTkEntry(self.add_wine_lib_frame, placeholder_text="Quantity")
        self.qt_entry_lib.place(x=120, y=200)
        self.qt_entry_lib.insert(0, "0")

        # Input Notes
        self.new_notes_lib = customtkinter.CTkLabel(self.add_wine_lib_frame, text="Notes:",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.new_notes_lib.place(x=20, y=250)

        self.notes_entry_lib = customtkinter.CTkEntry(self.add_wine_lib_frame, placeholder_text="Notes")
        self.notes_entry_lib.place(x=120, y=250)
        self.notes_entry_lib.insert(0, "0")

        # Confirm Button
        # self.confirm_btn_lib = customtkinter.CTkButton(self.add_wine_lib_frame, text="Confirm", command=self.confirm_add)
        # elf.confirm_btn_lib.place(x=20, y=450)

    def sch_screen(self):
        self.sch_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.sch_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")
        # self.main_frame.grid_rowconfigure(6, weight=1)

        # Schedule label
        self.sch_label = customtkinter.CTkLabel(self.sch_frame, text="Schedule",
                                                font=customtkinter.CTkFont(size=30, weight="bold"))
        self.sch_label.place(x=20, y=20)
        # Set up start date combo boxes
        today = date.today()
        self.start_label = customtkinter.CTkLabel(self.sch_frame, text="Start date",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.start_label.place(x=20, y=70)
        self.start_month_drop = customtkinter.CTkComboBox(self.sch_frame,
                                                          values=["01", "02", "03", "04", "05", "06", "07",
                                                                  "08", "09", "10", "11", "12"])
        self.start_month_drop.place(x=20, y=100)
        self.start_month_drop.set(today.month)
        self.start_day_drop = customtkinter.CTkComboBox(self.sch_frame, values=[str(i + 1) for i in range(31)])
        self.start_day_drop.place(x=170, y=100)
        self.start_day_drop.set(today.day)
        self.start_year_drop = customtkinter.CTkComboBox(self.sch_frame, values=[str(i + 2023) for i in range(15)])
        self.start_year_drop.place(x=320, y=100)
        self.start_year_drop.set(today.year)

        # Set up end date combo boxes
        self.end_label = customtkinter.CTkLabel(self.sch_frame, text="End date",
                                                font=customtkinter.CTkFont(size=20, weight="bold"))
        self.end_label.place(x=20, y=170)
        self.end_month_drop = customtkinter.CTkComboBox(self.sch_frame,
                                                        values=["01", "02", "03", "04", "05", "06", "07",
                                                                "08", "09", "10", "11", "12"])
        self.end_month_drop.place(x=20, y=200)
        self.end_month_drop.set(today.month)  # (str((int(today.month)+3) % 12))
        self.end_day_drop = customtkinter.CTkComboBox(self.sch_frame, values=[str(i + 1) for i in range(31)])
        self.end_day_drop.place(x=170, y=200)
        self.end_day_drop.set(today.day)
        self.end_year_drop = customtkinter.CTkComboBox(self.sch_frame, values=[str(i + 2023) for i in range(15)])
        self.end_year_drop.place(x=320, y=200)
        self.end_year_drop.set(today.year)

        self.generate_sch = customtkinter.CTkButton(self.sch_frame, text="Generate Schedule", command=self.gen_sch)
        self.generate_sch.place(x=20, y=300)

    def gen_sch(self):
        temp_var = self.get_maximum_rows(sheet_object=current) - 1
        self.start_date = date(int(self.start_year_drop.get()), int(self.start_month_drop.get()),
                               int(self.start_day_drop.get()))
        self.end_date = date(int(self.end_year_drop.get()), int(self.end_month_drop.get()),
                             int(self.end_day_drop.get()))
        overall_index = 0
        # Generate first racking
        self.first_rack = self.add_date(14)
        for i in range(temp_var):
            sch["A" + str(i + 1)] = str(self.first_rack.year) + "-" + str(self.first_rack.month) + "-" + (
                str(int(self.first_rack.day) + i))
            sch["B" + str(i + 1)] = "First racking of " + current["B" + str(i + 2)].value
            overall_index += 1
        # Generate Second racking
        self.second_rack = self.add_date(28)
        for i in range(temp_var):
            sch["A" + str(overall_index + 1)] = str(self.second_rack.year) + "-" + str(self.second_rack.month) + "-" + (
                str(int(self.second_rack.day) + i))
            sch["B" + str(overall_index + 1)] = "Second racking of " + current["B" + str(i + 2)].value
            overall_index += 1

        # Generate Bentonite Addition
        self.add_bent = self.sub_date(30)
        for i in range(temp_var):
            sch["A" + str(overall_index + 1)] = str(self.add_bent.year) + "-" + str(self.add_bent.month) + "-" + (
                str(int(self.add_bent.day) + i))
            sch["B" + str(overall_index + 1)] = "Bentonite Addition of " + current["B" + str(i + 2)].value
            overall_index += 1

        # Generate Bentonite Addition
        self.rack_bent = self.sub_date(23)
        for i in range(temp_var):
            sch["A" + str(overall_index + 1)] = str(self.rack_bent.year) + "-" + str(self.rack_bent.month) + "-" + (
                str(int(self.rack_bent.day) + i))
            sch["B" + str(overall_index + 1)] = "Bentonite Racking of " + current["B" + str(i + 2)].value
            overall_index += 1

        # Generate rough filtering
        self.rough_filter = self.sub_date(14)
        for i in range(temp_var):
            sch["A" + str(overall_index + 1)] = str(self.rough_filter.year) + "-" + str(
                self.rough_filter.month) + "-" + (str(int(self.rough_filter.day) + i))
            sch["B" + str(overall_index + 1)] = "Rough Filtering of " + current["B" + str(i + 2)].value
            overall_index += 1

        # Generate sweetening
        self.sweet = self.sub_date(8)
        for i in range(temp_var):
            sch["A" + str(overall_index + 1)] = str(self.sweet.year) + "-" + str(self.sweet.month) + "-" + (
                str(int(self.sweet.day)))
            sch["B" + str(overall_index + 1)] = "Sweetening of " + current["B" + str(i + 2)].value
            overall_index += 1

        # Generate Final filtering
        self.final_filter = self.sub_date(3)
        for i in range(temp_var):
            sch["A" + str(overall_index + 1)] = str(self.final_filter.year) + "-" + str(
                self.final_filter.month) + "-" + (str(int(self.final_filter.day)))
            sch["B" + str(overall_index + 1)] = "Final Filtering of " + current["B" + str(i + 2)].value
            overall_index += 1

        book.save("WineMakerData.xlsx")

        self.schedule_page()

    def schedule_page(self):
        self.sch_page_frame = customtkinter.CTkFrame(self, width=120, corner_radius=0)
        self.sch_page_frame.grid(row=0, column=1, rowspan=4, sticky="nsew")
        # self.main_frame.grid_rowconfigure(6, weight=1)

        # Schedule label
        self.sch_page_label = customtkinter.CTkLabel(self.sch_page_frame, text="Schedule",
                                                     font=customtkinter.CTkFont(size=30, weight="bold"))
        self.sch_page_label.place(x=0, y=20)

        # New Schedule button
        self.new_sch = customtkinter.CTkButton(self.sch_page_frame, text="New Schedule", command=self.sch_screen)
        self.new_sch.place(x=550, y=20)

        # Scroll frame
        self.scrollable_sch = customtkinter.CTkScrollableFrame(self.sch_page_frame, width=700, height=450)
        self.scrollable_sch.place(x=0, y=70)
        # Generate schedule layout
        for i in range(self.get_maximum_rows(sheet_object=sch)):
            sch_date = customtkinter.CTkLabel(self.scrollable_sch, text=sch["A" + str(i + 1)].value + ":",
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
            sch_date.grid(row=i, column=0, padx=0, pady=(0, 20))

            self.scrollable_sch.columnconfigure(1, weight=1)

            sch_data = customtkinter.CTkLabel(self.scrollable_sch, text=sch["B" + str(i + 1)].value,
                                              font=customtkinter.CTkFont(size=20, weight="bold"))
            sch_data.grid(row=i, column=2, padx=150, pady=(0, 20))

        # Returns to create new schedule screen if no schedule exists
        if (sch["A1"].value == None):
            self.sch_screen()

    def add_date(self, added):  # This function adds "added" days to the start date for scheduling
        new_day = int(self.start_day_drop.get()) + added
        new_month = int(self.start_month_drop.get())
        new_year = int(self.start_year_drop.get())
        if (new_month == 1):
            if (new_day > 31):
                new_day = new_day - 31
                new_month = new_month + 1

        if (new_month == 2 and new_year % 4):
            if (new_day > 29):
                new_day = new_day - 29
                new_month = new_month + 1
        elif (new_month == 2):
            if (new_day > 28):
                new_day = new_day - 28
                new_month = new_month + 1

        if (new_month == 3):
            if (new_day > 31):
                new_day = new_day - 31
                new_month = new_month + 1

        if (new_month == 4):
            if (new_day > 30):
                new_day = new_day - 30
                new_month = new_month + 1

        if (new_month == 5):
            if (new_day > 31):
                new_day = new_day - 31
                new_month = new_month + 1

        if (new_month == 6):
            if (new_day > 30):
                new_day = new_day - 30
                new_month = new_month + 1

        if (new_month == 7):
            if (new_day > 31):
                new_day = new_day - 31
                new_month = new_month + 1

        if (new_month == 8):
            if (new_day > 31):
                new_day = new_day - 31
                new_month = new_month + 1

        if (new_month == 9):
            if (new_day > 30):
                new_day = new_day - 30
                new_month = new_month + 1

        if (new_month == 10):
            if (new_day > 31):
                new_day = new_day - 31
                new_month = new_month + 1

        if (new_month == 11):
            if (new_day > 30):
                new_day = new_day - 30
                new_month = new_month + 1

        if (new_month == 12):
            if (new_day > 31):
                new_day = new_day - 31
                new_month = 1
                new_year = new_year + 1

        return date(new_year, new_month, new_day)

    def sub_date(self, sub):  # This function adds "added" days to the start date for scheduling
        new_day = int(self.end_day_drop.get())
        new_month = int(self.end_month_drop.get())
        new_year = int(self.end_year_drop.get())

        while (sub > 0):

            if ((new_day - sub) >= 0):
                new_day = new_day - sub
                sub = 0
            else:
                if (new_month == 12):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 30

                elif (new_month == 11):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 31

                elif (new_month == 10):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 30

                elif (new_month == 9):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 31

                elif (new_month == 8):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 31

                elif (new_month == 7):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 30

                elif (new_month == 6):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 31

                elif (new_month == 5):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 30

                elif (new_month == 4):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 31

                elif (new_month == 3):
                    sub = sub - new_day
                    new_month = new_month - 1
                    if (new_month == 2 and new_year % 4):
                        new_day = 29
                    elif (new_month == 2):
                        new_day = 28

                elif (new_month == 2):
                    sub = sub - new_day
                    new_month = new_month - 1
                    new_day = 31

                elif (new_month == 1):
                    sub = sub - new_day
                    new_month = 12
                    new_day = 31
                    new_year = new_year - 1

        return date(new_year, new_month, new_day)

    def get_maximum_rows(self, sheet_object):
        rows = 0
        for max_row, row in enumerate(sheet_object, 1):
            if not all(col.value is None for col in row):
                rows += 1
        return rows

    def round_half_up(slef, n, decimals=2):
        multiplier = 10 ** decimals
        return math.floor(n * multiplier + 0.5) / multiplier


if __name__ == "__main__":
    # root = customtkinter.CTk()
    main = App()
    main.mainloop()
